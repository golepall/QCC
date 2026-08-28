"""Excel模板文件管理服务 — 扫描、预览、写入Excel测试报告模板"""
import os
import sys
from pathlib import Path
from datetime import datetime

# 模板文件目录
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "docs", "reference", "测试报告文件")
TEMPLATE_DIR = os.path.normpath(TEMPLATE_DIR)


def get_template_dir() -> str:
    """获取模板目录的绝对路径，检查目录是否存在和可访问"""
    abs_path = os.path.abspath(TEMPLATE_DIR)
    if not os.path.isdir(abs_path):
        raise FileNotFoundError(f"模板目录不存在: {abs_path}")
    if not os.access(abs_path, os.R_OK):
        raise PermissionError(f"无读取权限: {abs_path}")
    return abs_path


def list_excel_files() -> list[dict]:
    """扫描模板目录，返回所有xlsx文件列表"""
    dir_path = get_template_dir()
    files = []
    for f in os.listdir(dir_path):
        if f.lower().endswith(".xlsx") and not f.startswith("~$"):
            file_path = os.path.join(dir_path, f)
            stat = os.stat(file_path)
            files.append({
                "filename": f,
                "file_path": file_path,
                "file_size": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "readable": os.access(file_path, os.R_OK),
                "writable": os.access(file_path, os.W_OK),
            })
    files.sort(key=lambda x: x["filename"])
    return files


def read_excel_preview(filename: str, max_rows: int = 50) -> dict:
    """读取Excel文件内容，返回各sheet的表头和数据预览"""
    import openpyxl

    dir_path = get_template_dir()
    file_path = os.path.join(dir_path, filename)

    # 安全检查：防止路径穿越
    if not os.path.abspath(file_path).startswith(os.path.abspath(dir_path)):
        raise PermissionError("非法文件路径")
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件不存在: {filename}")
    if not os.access(file_path, os.R_OK):
        raise PermissionError(f"无读取权限: {filename}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # 跳过全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if i == 0:
                headers = row_data
            elif i <= max_rows:
                rows.append(row_data)
            else:
                break

        sheets_data.append({
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows,
            "total_rows": ws.max_row or 0,
            "total_cols": ws.max_column or 0,
        })

    wb.close()
    return {
        "filename": filename,
        "sheets": sheets_data,
    }


def write_data_to_template(filename: str, sheet_name: str, data: list[dict], start_row: int = 2) -> dict:
    """将业务数据写入Excel模板的指定sheet

    Args:
        filename: Excel文件名
        sheet_name: 目标sheet名称
        data: 要写入的数据，每行为一个dict，key为列名
        start_row: 起始行号（默认从第2行开始，第1行为表头）
    """
    import openpyxl

    dir_path = get_template_dir()
    file_path = os.path.join(dir_path, filename)

    # 安全检查
    if not os.path.abspath(file_path).startswith(os.path.abspath(dir_path)):
        raise PermissionError("非法文件路径")
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件不存在: {filename}")
    if not os.access(file_path, os.W_OK):
        raise PermissionError(f"无写入权限: {filename}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}")

    ws = wb[sheet_name]

    # 读取表头，建立列名到列号的映射
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header_map[str(cell.value).strip()] = col_idx

    # 写入数据
    written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_name, value in row_data.items():
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx, value=value)
                written += 1

    wb.save(file_path)
    wb.close()

    return {
        "filename": filename,
        "sheet_name": sheet_name,
        "rows_written": len(data),
        "cells_written": written,
    }


def check_permissions() -> dict:
    """检查模板目录的读写权限"""
    try:
        dir_path = get_template_dir()
        readable = os.access(dir_path, os.R_OK)
        writable = os.access(dir_path, os.W_OK)

        # 检查子文件
        xlsx_files = [f for f in os.listdir(dir_path) if f.lower().endswith(".xlsx")]
        file_permissions = []
        for f in xlsx_files[:5]:  # 只检查前5个
            fp = os.path.join(dir_path, f)
            file_permissions.append({
                "filename": f,
                "readable": os.access(fp, os.R_OK),
                "writable": os.access(fp, os.W_OK),
            })

        return {
            "dir_path": dir_path,
            "dir_exists": True,
            "dir_readable": readable,
            "dir_writable": writable,
            "xlsx_count": len(xlsx_files),
            "sample_files": file_permissions,
        }
    except Exception as e:
        return {
            "dir_path": TEMPLATE_DIR,
            "dir_exists": False,
            "error": str(e),
        }
