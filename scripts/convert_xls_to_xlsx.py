"""将所有旧版.xls文件转换为新版.xlsx格式"""
import os
import xlrd
from openpyxl import Workbook

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "docs", "reference", "测试报告文件")
TEMPLATE_DIR = os.path.normpath(TEMPLATE_DIR)


def convert_xls_to_xlsx(xls_path: str) -> str:
    """将单个.xls文件转换为.xlsx格式，返回新文件路径"""
    # 打开.xls文件
    xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
    
    # 创建新的.xlsx工作簿
    xlsx_book = Workbook()
    # 删除默认创建的Sheet
    xlsx_book.remove(xlsx_book.active)
    
    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        xlsx_sheet = xlsx_book.create_sheet(title=sheet_name)
        
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = cell.value
                
                # 处理日期类型
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(value, xls_book.datemode)
                        from datetime import datetime
                        value = datetime(*date_tuple)
                    except:
                        pass
                
                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
    
    # 生成新的文件路径（.xls -> .xlsx）
    xlsx_path = xls_path + "x"  # .xls + x = .xlsx
    xlsx_book.save(xlsx_path)
    return xlsx_path


def convert_all():
    """转换目录下所有.xls文件"""
    if not os.path.isdir(TEMPLATE_DIR):
        print(f"目录不存在: {TEMPLATE_DIR}")
        return
    
    xls_files = [f for f in os.listdir(TEMPLATE_DIR) if f.lower().endswith(".xls") and not f.lower().endswith(".xlsx")]
    
    if not xls_files:
        print("没有找到.xls文件")
        return
    
    print(f"找到 {len(xls_files)} 个.xls文件，开始转换...\n")
    
    success = 0
    failed = 0
    
    for xls_file in xls_files:
        xls_path = os.path.join(TEMPLATE_DIR, xls_file)
        xlsx_path = xls_path + "x"
        
        # 检查是否已存在对应的.xlsx文件
        if os.path.exists(xlsx_path):
            print(f"[跳过] {xls_file} -> 已存在对应的.xlsx文件")
            success += 1
            continue
        
        try:
            new_path = convert_xls_to_xlsx(xls_path)
            file_size = os.path.getsize(new_path)
            print(f"[成功] {xls_file} -> {os.path.basename(new_path)} ({file_size / 1024:.1f} KB)")
            success += 1
        except Exception as e:
            print(f"[失败] {xls_file} -> 错误: {e}")
            failed += 1
    
    print(f"\n转换完成: 成功 {success} 个, 失败 {failed} 个")
    
    # 列出转换后的文件
    print(f"\n当前目录下所有xlsx文件:")
    xlsx_files = [f for f in os.listdir(TEMPLATE_DIR) if f.lower().endswith(".xlsx")]
    for i, f in enumerate(sorted(xlsx_files), 1):
        print(f"  {i}. {f}")


if __name__ == "__main__":
    convert_all()
