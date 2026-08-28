"""Excel 报告生成服务（openpyxl 实现）"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.template import TestCategory, TestItem
from app.models.project import TestRecord


async def generate_report(db: AsyncSession, project, template) -> str:
    """生成测试报告 Excel 文件，返回文件路径"""
    wb = Workbook()
    # 移除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 封面
    ws_cover = wb.create_sheet("封面", 0)
    _write_cover(ws_cover, project, template)

    # 获取全部分类和测试项
    cat_result = await db.execute(
        select(TestCategory).where(TestCategory.template_id == project.template_id).order_by(TestCategory.sort_order)
    )
    categories = list(cat_result.scalars().all())

    for cat in categories:
        item_result = await db.execute(
            select(TestItem).where(TestItem.category_id == cat.id).order_by(TestItem.sort_order)
        )
        items = list(item_result.scalars().all())

        # 获取记录
        rec_result = await db.execute(
            select(TestRecord).where(
                TestRecord.project_id == project.id,
                TestRecord.item_id.in_([i.id for i in items])
            )
        )
        record_map = {r.item_id: r for r in rec_result.scalars().all()}

        ws = wb.create_sheet(cat.sheet_name or cat.category_name)
        _write_category_sheet(ws, cat, items, record_map)

    # 保存
    export_dir = _ensure_dir()
    file_path = os.path.join(export_dir, f"{project.project_code}_report.xlsx")
    wb.save(file_path)
    return file_path


def _write_cover(ws, project, template):
    """写封面"""
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"{template.name if template else ''} — {project.product_name}"
    c.font = Font(size=16, bold=True)
    c.alignment = Alignment(horizontal="center")

    info = [
        ("项目编号", project.project_code),
        ("产品型号", project.product_model),
        ("测试类型", project.test_type),
        ("测试人员", project.tester),
        ("状态", project.status),
    ]
    for i, (label, value) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(value or ""))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30


def _write_category_sheet(ws, cat, items, record_map):
    """写分类 sheet"""
    headers = ["编号", "测试项目", "测试步骤/条件", "判定标准", "结果", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    row = 2
    for item in items:
        if item.is_header:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1, value=f"{item.item_no} {item.test_item}")
            c.font = Font(bold=True, size=11)
            c.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            row += 1
            continue

        record = record_map.get(item.id)
        data = [
            item.item_no, item.test_item, item.test_case or item.condition_desc or "",
            item.criteria or "", record.result if record else "", record.comment if record else "",
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=str(val))
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 25


def _ensure_dir() -> str:
    d = os.path.join(os.path.dirname(__file__), "..", "..", "..", "server", "exports")
    os.makedirs(d, exist_ok=True)
    return d
